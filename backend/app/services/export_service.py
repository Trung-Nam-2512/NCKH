"""
Export Service - Export kết quả phân tích ra PDF, Excel, PNG
"""
import io
import base64
from typing import Dict, Any, List, Optional
import logging
from datetime import datetime
import pandas as pd

# PDF and Excel dependencies
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.lineplots import LinePlot
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logging.warning("ReportLab not available - PDF export disabled")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("OpenPyXL not available - Excel export disabled")

class ExportService:
    """Service để export kết quả phân tích ra các định dạng khác nhau"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet() if REPORTLAB_AVAILABLE else None
    
    def export_to_excel(self, comprehensive_result: Dict[str, Any], filename: str = None) -> bytes:
        """Export kết quả phân tích ra file Excel"""
        
        if not OPENPYXL_AVAILABLE:
            raise ValueError("OpenPyXL not installed - cannot export to Excel")
        
        try:
            # Tạo workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # 1. Summary Sheet
            ws_summary = wb.create_sheet("Tóm tắt phân tích")
            self._create_summary_sheet(ws_summary, comprehensive_result)
            
            # 2. Distribution Comparison Sheet
            ws_dist = wb.create_sheet("So sánh phân phối")
            self._create_distribution_comparison_sheet(ws_dist, comprehensive_result)
            
            # 3. Return Periods Sheet
            ws_return = wb.create_sheet("Chu kỳ lặp lại")
            self._create_return_periods_sheet(ws_return, comprehensive_result)
            
            # 4. Frequency Table Sheet
            if 'frequency_by_best_model' in comprehensive_result.get('frequency_analysis', {}):
                ws_freq = wb.create_sheet("Bảng tần suất")
                self._create_frequency_table_sheet(ws_freq, comprehensive_result)
            
            # 5. Raw Data Sheet
            if 'data_summary' in comprehensive_result:
                ws_data = wb.create_sheet("Dữ liệu gốc")
                self._create_raw_data_sheet(ws_data, comprehensive_result)
            
            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            logging.error(f"Error exporting to Excel: {e}")
            raise
    
    def export_to_pdf(self, comprehensive_result: Dict[str, Any], filename: str = None) -> bytes:
        """Export kết quả phân tích ra file PDF"""
        
        if not REPORTLAB_AVAILABLE:
            raise ValueError("ReportLab not installed - cannot export to PDF")
        
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=self.styles['Title'],
                fontSize=18,
                textColor=colors.darkblue,
                spaceAfter=30
            )
            
            story.append(Paragraph("BÁO CÁO PHÂN TÍCH TẦN SUẤT THỦY VĂN", title_style))
            story.append(Spacer(1, 20))
            
            # Metadata
            if 'metadata' in comprehensive_result:
                metadata = comprehensive_result['metadata']
                story.append(Paragraph(f"<b>Ngày phân tích:</b> {metadata.get('analysis_date', 'N/A')}", self.styles['Normal']))
                story.append(Paragraph(f"<b>Loại phân tích:</b> {metadata.get('analysis_type', 'N/A')}", self.styles['Normal']))
                story.append(Paragraph(f"<b>Số năm dữ liệu:</b> {metadata.get('data_years', 'N/A')}", self.styles['Normal']))
                story.append(Paragraph(f"<b>Cấp độ phân tích:</b> {metadata.get('analysis_grade', 'N/A')}", self.styles['Normal']))
                story.append(Spacer(1, 20))
            
            # Best Distribution Summary
            if 'statistical_analysis' in comprehensive_result:
                best_dist = comprehensive_result['statistical_analysis']['statistical_summary']['best_distribution']
                
                story.append(Paragraph("1. PHÂN PHỐI TỐI ưU", self.styles['Heading2']))
                story.append(Paragraph(f"<b>Phân phối:</b> {best_dist['display_name']}", self.styles['Normal']))
                story.append(Paragraph(f"<b>AIC:</b> {best_dist['aic']:.2f}", self.styles['Normal']))
                story.append(Paragraph(f"<b>P-value:</b> {best_dist.get('p_value', 'N/A')}", self.styles['Normal']))
                story.append(Paragraph(f"<b>Chất lượng dữ liệu:</b> {best_dist.get('data_quality_grade', 'N/A')}", self.styles['Normal']))
                story.append(Spacer(1, 20))
            
            # Return Periods Table
            if 'frequency_analysis' in comprehensive_result and 'return_periods_analysis' in comprehensive_result['frequency_analysis']:
                story.append(Paragraph("2. CHU KỲ LẶP LẠI", self.styles['Heading2']))
                
                return_periods = comprehensive_result['frequency_analysis']['return_periods_analysis']
                table_data = [['Chu kỳ (năm)', 'Tần suất (%)', 'Lưu lượng (m³/s)', 'Xác suất vượt quá']]
                
                for rp in return_periods:
                    table_data.append([
                        str(rp['return_period']),
                        f"{rp['frequency_percent']:.2f}%",
                        str(rp['discharge_value']),
                        f"{rp['exceedance_probability']:.4f}"
                    ])
                
                table = Table(table_data, colWidths=[1.5*inch, 1.5*inch, 2*inch, 2*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(table)
                story.append(Spacer(1, 20))
            
            # Distribution Comparison
            if 'statistical_analysis' in comprehensive_result:
                goodness_of_fit = comprehensive_result['statistical_analysis']['statistical_summary']['goodness_of_fit_ranking']
                
                story.append(Paragraph("3. SO SÁNH CÁC PHÂN PHỐI", self.styles['Heading2']))
                
                table_data = [['Thứ hạng', 'Phân phối', 'AIC', 'P-value', 'Đánh giá']]
                for item in goodness_of_fit[:5]:  # Top 5
                    table_data.append([
                        str(item['rank']),
                        item['distribution'].capitalize(),
                        f"{item['aic']:.2f}",
                        f"{item['p_value']:.4f}" if item['p_value'] else "N/A",
                        item['status']
                    ])
                
                table = Table(table_data, colWidths=[1*inch, 1.5*inch, 1*inch, 1*inch, 2.5*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(table)
                story.append(Spacer(1, 20))
            
            # Recommendations
            if 'metadata' in comprehensive_result and 'recommendations' in comprehensive_result['metadata']:
                story.append(Paragraph("4. KHUYẾN NGHỊ", self.styles['Heading2']))
                
                for i, rec in enumerate(comprehensive_result['metadata']['recommendations'], 1):
                    story.append(Paragraph(f"{i}. {rec}", self.styles['Normal']))
                
                story.append(Spacer(1, 20))
            
            # Add charts if available
            if 'visualizations' in comprehensive_result:
                visualizations = comprehensive_result['visualizations']
                
                # Frequency curve
                if 'frequency_curve' in visualizations and visualizations['frequency_curve'].get('plot_base64'):
                    story.append(Paragraph("5. ĐƯỜNG CONG TẦN SUẤT", self.styles['Heading2']))
                    
                    # Convert base64 to image
                    plot_data = visualizations['frequency_curve']['plot_base64']
                    if plot_data.startswith('data:image/png;base64,'):
                        plot_data = plot_data.replace('data:image/png;base64,', '')
                    
                    img_buffer = io.BytesIO(base64.b64decode(plot_data))
                    img = Image(img_buffer, width=6*inch, height=4*inch)
                    story.append(img)
                    story.append(Spacer(1, 20))
            
            # Build PDF
            doc.build(story)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            logging.error(f"Error exporting to PDF: {e}")
            raise
    
    def _create_summary_sheet(self, ws, comprehensive_result):
        """Tạo sheet tóm tắt"""
        
        # Title
        ws['A1'] = "BÁO CÁO PHÂN TÍCH TẦN SUẤT THỦY VĂN"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Metadata
        row = 3
        if 'metadata' in comprehensive_result:
            metadata = comprehensive_result['metadata']
            
            ws[f'A{row}'] = "Thông tin phân tích"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            info_items = [
                ("Ngày phân tích", metadata.get('analysis_date', 'N/A')),
                ("Loại phân tích", metadata.get('analysis_type', 'N/A')),
                ("Số năm dữ liệu", metadata.get('data_years', 'N/A')),
                ("Tổng số bản ghi", metadata.get('total_records', 'N/A')),
                ("Cấp độ phân tích", metadata.get('analysis_grade', 'N/A'))
            ]
            
            for label, value in info_items:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1
        
        # Best distribution
        if 'statistical_analysis' in comprehensive_result:
            best_dist = comprehensive_result['statistical_analysis']['statistical_summary']['best_distribution']
            
            ws[f'A{row + 1}'] = "Phân phối tối ưu"
            ws[f'A{row + 1}'].font = Font(bold=True)
            row += 2
            
            dist_items = [
                ("Tên phân phối", best_dist['display_name']),
                ("AIC", f"{best_dist['aic']:.2f}"),
                ("P-value", best_dist.get('p_value', 'N/A')),
                ("Chất lượng dữ liệu", best_dist.get('data_quality_grade', 'N/A')),
                ("Mức độ bất định", best_dist.get('uncertainty_level', 'N/A'))
            ]
            
            for label, value in dist_items:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1
    
    def _create_distribution_comparison_sheet(self, ws, comprehensive_result):
        """Tạo sheet so sánh phân phối"""
        
        if 'statistical_analysis' not in comprehensive_result:
            return
        
        # Headers
        headers = ['Thứ hạng', 'Phân phối', 'AIC', 'P-value', 'Đánh giá']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        goodness_of_fit = comprehensive_result['statistical_analysis']['statistical_summary']['goodness_of_fit_ranking']
        
        for row, item in enumerate(goodness_of_fit, 2):
            ws.cell(row=row, column=1, value=item['rank'])
            ws.cell(row=row, column=2, value=item['distribution'].capitalize())
            ws.cell(row=row, column=3, value=f"{item['aic']:.2f}")
            ws.cell(row=row, column=4, value=f"{item['p_value']:.4f}" if item['p_value'] else "N/A")
            ws.cell(row=row, column=5, value=item['status'])
    
    def _create_return_periods_sheet(self, ws, comprehensive_result):
        """Tạo sheet chu kỳ lặp lại"""
        
        if 'frequency_analysis' not in comprehensive_result or 'return_periods_analysis' not in comprehensive_result['frequency_analysis']:
            return
        
        # Headers
        headers = ['Chu kỳ (năm)', 'Tần suất (%)', 'Lưu lượng (m³/s)', 'Xác suất vượt quá']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        return_periods = comprehensive_result['frequency_analysis']['return_periods_analysis']
        
        for row, rp in enumerate(return_periods, 2):
            ws.cell(row=row, column=1, value=rp['return_period'])
            ws.cell(row=row, column=2, value=f"{rp['frequency_percent']:.2f}")
            ws.cell(row=row, column=3, value=rp['discharge_value'])
            ws.cell(row=row, column=4, value=f"{rp['exceedance_probability']:.4f}")
    
    def _create_frequency_table_sheet(self, ws, comprehensive_result):
        """Tạo sheet bảng tần suất"""
        
        freq_data = comprehensive_result['frequency_analysis']['frequency_by_best_model']
        
        if not freq_data.get('theoretical_curve'):
            return
        
        # Headers
        headers = ['STT', 'Tần suất (%)', 'Lưu lượng (m³/s)', 'Thời gian lặp lại (năm)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data (first 20 rows to avoid too large sheet)
        theoretical_curve = freq_data['theoretical_curve'][:20]
        
        for row, point in enumerate(theoretical_curve, 2):
            ws.cell(row=row, column=1, value=point.get('Thứ tự', row-1))
            ws.cell(row=row, column=2, value=point.get('Tần suất P(%)', 'N/A'))
            ws.cell(row=row, column=3, value=point.get('Lưu lượng dòng chảy Q m³/s', 'N/A'))
            ws.cell(row=row, column=4, value=point.get('Thời gian lặp lại (năm)', 'N/A'))
    
    def _create_raw_data_sheet(self, ws, comprehensive_result):
        """Tạo sheet dữ liệu gốc"""
        
        if 'data_summary' not in comprehensive_result or 'annual_data' not in comprehensive_result['data_summary']:
            return
        
        annual_data = comprehensive_result['data_summary']['annual_data']
        
        # Headers
        headers = ['Năm', 'Min', 'Max', 'Mean', 'Std', 'Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        for row, (year, data) in enumerate(annual_data.items(), 2):
            ws.cell(row=row, column=1, value=year)
            ws.cell(row=row, column=2, value=data.get('min', 'N/A'))
            ws.cell(row=row, column=3, value=data.get('max', 'N/A'))
            ws.cell(row=row, column=4, value=data.get('mean', 'N/A'))
            ws.cell(row=row, column=5, value=data.get('std', 'N/A'))
            ws.cell(row=row, column=6, value=data.get('count', 'N/A'))